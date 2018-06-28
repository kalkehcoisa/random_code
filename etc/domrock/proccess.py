#!/usr/bin/env python3

from openpyxl import load_workbook
import pandas as pd


def pre_gera_xls(input_file, output_file, low_memory=True):
    '''
    Gera a planilha base para calcular os saldos.
    '''
    movto_file = pd.ExcelFile(
        input_file,
        low_memory=low_memory,
        dtype={
            'item': 'str',
            'tipo_movimento': 'str',
            'data_lancamento': 'date',
            'quantidade': 'float',
            'valor': 'float'
        },
    )
    movto_data = movto_file.parse(movto_file.sheet_names[0])

    result = movto_data.groupby(
        ['item', 'tipo_movimento', 'data_lancamento']).sum()\
        .reset_index().sort_values(by=['data_lancamento', 'item'])[1:]

    result.to_excel(output_file, index=False)


def recupera_saldos_base(input_file, output_file, low_memory=True):
    '''
    Recupera os saldos iniciais para fazer os
    cálculos dos saldos no arquivo pré processado.
    '''

    saldo_file = pd.ExcelFile(
        input_file,
        low_memory=low_memory,
        dtype={
            'item': 'str',
            'data_inicio': 'date'
        }
    )
    saldo_data = saldo_file.parse(saldo_file.sheet_names[0])
    columns = ['item', 'data_inicio', 'qtd_inicio', 'valor_inicio']
    saldos = {}
    for item in saldo_data[columns].to_dict(orient='records'):
        saldos[item['item']] = {
            'data': item['data_inicio'].to_pydatetime(),
            'quantidade': item['qtd_inicio'],
            'valor': item['valor_inicio'],
        }
    return saldos


if __name__ == '__main__':
    input_file = 'SaldoITEM.XLSX'
    output_file = 'output.xlsx'

    pre_gera_xls(input_file='MovtoITEM.XLSX', output_file=output_file)
    saldos = recupera_saldos_base(
        input_file=input_file, output_file=output_file)

    colunas = (
        'item', 'tipo_movimento', 'data_lancamento',
        'quantidade', 'valor', 'saldo_inicial_qt', 'saldo_inicial_valor',
        'saldo_final_qt', 'saldo_final_valor'
    )

    wb = load_workbook(filename=output_file)
    ws = wb.active
    for i, col in enumerate(colunas, 1):
        ws.cell(row=1, column=i).value = col

    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=9), 2):
        linha = dict(zip(colunas, [p.value for p in row]))
        saldo = saldos[linha['item']]

        ws.cell(row=i, column=3).number_format = 'D/M/YYYY'
        for f_col in range(4, 10):
            ws.cell(row=i, column=f_col).number_format = '0.00'

        # saldo_inicial_qt - saldo_inicial_valor
        ws.cell(row=i, column=6).value = saldo['quantidade']
        ws.cell(row=i, column=7).value = saldo['valor']

        # saldo_final_qt - saldo_final_valor
        if linha['tipo_movimento'] == 'Ent':
            saldo['quantidade'] += float(ws.cell(row=i, column=4).value or 0)
            saldo['valor'] += float(ws.cell(row=i, column=5).value or 0)
        elif linha['tipo_movimento'] == 'Sai':
            saldo['quantidade'] -= float(ws.cell(row=i, column=4).value or 0)
            saldo['valor'] -= float(ws.cell(row=i, column=5).value or 0)
        ws.cell(row=i, column=8).value = saldo['quantidade']
        ws.cell(row=i, column=9).value = saldo['valor']

    # trata as larguras das colunas
    dims = {}
    for row in ws.rows:
        for cell in row:
            dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    wb.save(filename=output_file)
